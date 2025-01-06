import os.path as osp
import logging
import yaml
from torch.utils.tensorboard import SummaryWriter
import torch
import os
import shutil
import pandas as pd
import numpy as np
import openpyxl
import random
import matplotlib.pyplot as plt


# Log File Generator
def create_exp_dir(dataset, model, name):
    '''
    日志文件存储地址
    :param dataset:
    :param model:
    :param name:
    :return:
    '''
    exp_dir = osp.join('exps', dataset, model, name)
    os.makedirs(exp_dir, exist_ok=True)
    return exp_dir


# Write to excel
def append_df_to_excel(filename, df):
    key = list(df.keys())
    row = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
           'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN']
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        for i in range(len(key)):
            sheet[row[i] + '1'] = key[i]
    else:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
    row = sheet.max_row + 1
    for i in range(len(key)):
        sheet.cell(row=row, column=i + 1, value=df[key[i]])
    wb.save(filename)
    wb.close()


# Log file
def get_logger(exp_dir):
    logger = logging.getLogger()

    logger.setLevel(logging.INFO)

    hdlr = logging.FileHandler(osp.join(exp_dir, 'log.txt'))
    console = logging.StreamHandler()
    fmtr = logging.Formatter('%(message)s')
    hdlr.setFormatter(fmtr)
    console.setFormatter(fmtr)

    logger.addHandler(hdlr)
    logger.addHandler(console)

    return logger


# Create tensorboard
def get_write(args):
    if os.path.exists(f'exps/{args["dataset_name"]}/{args["model_name"]}/{args["exp_name"]}/write') == True:
        shutil.rmtree(f'exps/{args["dataset_name"]}/{args["model_name"]}/{args["exp_name"]}/write')
    write = SummaryWriter(f'exps/{args["dataset_name"]}/{args["model_name"]}/{args["exp_name"]}/write')

    return write


# model_size
def model_size(model, type_size=4):
    size = 0
    for p in model.parameters():
        size += p.numel() * type_size  # Bytes

    return size


# normalize
def normalize(tensors, mean, std, type='zscore'):
    '''

    :param tensors: list
    :param mean:
    :param std:
    :param type:
    :return:
    '''
    y = []
    for x in tensors:
        if type == 'zscore':
            y.append((x - mean) / std)
        elif type == 'maxmin':
            _max, _min = mean, std
            z = (x - _min) / (_max - _min)
            z = z * 2 - 1  # [-1, 1]
            y.append(z)
        else:
            raise ValueError('type should be zscore or maxmin')

    return y


# denormalize
def denormalize(tensor, base):
    mean = base[0]
    std = base[1]
    y = (std * tensor + mean)
    return y


# move2device
def move2device(x, device):
    if isinstance(x, list):
        y = []
        for item in x:
            y.append(move2device(item, device))
    elif isinstance(x, dict):
        y = {}
        for k, v in x.items():
            y[k] = move2device(v, device)
    elif x is None:
        y = None
    else:
        y = x.to(device)

    return y


# totensor
def totensor(x):
    if isinstance(x, list):
        y = []
        for item in x:
            y.append(totensor(item))
    elif isinstance(x, dict):
        y = {}
        for k, v in x.items():
            y[k] = totensor(v)
    elif x is None:
        y = None
    else:
        y = torch.as_tensor(x, dtype=torch.float32).detach()
    return y


# save_metrics
def save_metrics(rmse, mae, mape1, file, mask=False):
    data = torch.stack((rmse, mae, mape1), dim=1)
    data_average = torch.mean(data, dim=0, keepdim=True)
    metrics = pd.DataFrame(data.numpy(), columns=['rmse', 'mae', 'mape'])
    metrics_average = pd.DataFrame(data_average.numpy(), columns=['rmse', 'mae', 'mape'])
    if mask == False:
        metrics.to_csv(osp.join(file, 'metrics.csv'), index=True)
        metrics_average.to_csv(osp.join(file, 'metrics_average.csv'), index=False)
    else:
        metrics.to_csv(osp.join(file, 'metrics_mask.csv'), index=True)
        metrics_average.to_csv(osp.join(file, 'metrics_average_mask.csv'), index=False)
    return metrics


# write2Yaml
def write2Yaml(data, save_path="test.yaml"):
    """
   存储yaml文件
   """
    with open(save_path, "w") as f:
        yaml.dump(data, f, sort_keys=False)


# Generate Record Dictionary
def hisdict(args):
    dic = {}
    dic['model_name'] = args['model_name']
    dic['dataset_name'] = args['dataset_name']
    dic['batch_size'] = args['batch_size']
    dic['max_epoch'] = args['epochs']
    dic['test'] = args['test']
    dic['debug'] = args['datasets']['debug']
    dic['in_len'] = args['datasets']['in_len']
    dic['out_len'] = args['datasets']['out_len']
    dic['features'] = " ".join(map(str, args['datasets']['features']))
    if args['model_name'] == 'AGMGRN' or args['model_name'] == 'AGMGT':
        dic['d_model'] = args['model']['d_model']
        dic['d_k'] = args['model']['d_k']
        dic['d_hidden_mt'] = args['model']['d_hidden_mt']
        dic['d_hidden_ff'] = args['model']['d_hidden_ff']
        dic['d_hidden_gm'] = args['model']['d_hidden_gm']
        dic['num_encoder_layers'] = args['model']['num_encoder_layers']
        dic['num_decoder_layers'] = args['model']['num_decoder_layers']
        dic['num_heads'] = args['model']['num_heads']
        dic['which_transition_matrices'] = " ".join(map(str, args['model']['which_transition_matrices']))
        aa = []
        if not args['model']['noTSA'] and not args['model']['noSSA'] and not args['model']['noML'] and not \
                args['model']['noAG'] and not args['model']['noTE'] and not args['model']['noSE'] and not \
                args['model']['noGC']:
            aa.append(0)
        if args['model']['noTSA']:
            aa.append(1)
        if args['model']['noSSA']:
            aa.append(2)
        if args['model']['noML']:
            aa.append(3)
        if args['model']['noAG']:
            aa.append(4)
        if args['model']['noGC']:
            aa.append(5)
        if args['model']['noTE']:
            aa.append(6)
        if args['model']['noSE']:
            aa.append(7)
        dic['no'] = " ".join(map(str, aa))
    else:
        dic['d_model'] = 'none'
        dic['d_k'] = 'none'
        dic['d_hidden_mt'] = 'none'
        dic['d_hidden_ff'] = 'none'
        dic['d_hidden_gm'] = 'none'
        dic['num_encoder_layers'] = 'none'
        dic['num_decoder_layers'] = 'none'
        dic['num_heads'] = 'none'
        dic['which_transition_matrices'] = 'none'
        dic['no'] = 'none'
    dic['epoch'] = -1
    dic['model_sizes'] = 'none'
    dic['gpu_memory'] = 'none'
    dic['infer_time'] = 'none'
    dic['rmse'] = 'none'
    dic['mae'] = 'none'
    dic['mape1'] = 'none'
    dic['mape2'] = 'none'
    dic['train_time'] = 'none'
    dic['add'] = 'none'

    return dic


# gpu memory print
def memory(id):
    print(id + " GPU memory allocated:", torch.cuda.memory_allocated() / 1024 / 1024 / 1024)
    print(id + " GPU memory cached:", torch.cuda.memory_cached() / 1024 / 1024 / 1024)


# Get Total Memory Usage
def get_total_cache_usage():
    total_allocated = 0.
    for i in range(torch.cuda.device_count()):
        device = torch.device(f'cuda:{i}')
        torch.cuda.set_device(device)
        allocated = torch.cuda.memory_reserved(device)
        total_allocated += allocated / (1024 ** 3)
    return total_allocated


def draw_result_Metro(num, inter, output, node_list, zero_data):
    output_data21 = output[:inter, num, node_list, 0].numpy()
    output_data21_new = np.zeros_like(output_data21)
    output_data22 = output[inter:2 * inter, num, node_list, 0].numpy()
    output_data22_new = np.zeros_like(output_data22)
    if num != 0:
        output_data21_new[num:] = output_data21[:-num]
        output_data22_new[num:] = output_data22[:-num]
    else:
        output_data21_new = output_data21
        output_data22_new = output_data22
    output_data2 = np.concatenate([output_data21_new, zero_data, output_data22_new], axis=0)
    return output_data2


def draw_result_Highway(num, inter, output, node_list):
    output_data = output[:inter, num, node_list, 0].numpy()
    output_data_new = np.zeros_like(output_data)
    if num != 0:
        output_data_new[num:] = output_data[:-num]
    else:
        output_data_new = output_data

    return output_data_new


def draw_result(dataset_name, data, num):
    output = data['output'][-576:]
    target = data['target'][-576:]
    node_list = [random.randint(0, output.shape[2] - 1) for _ in range(num)]

    if dataset_name == 'HZMetro' or dataset_name == 'SHMetro':
        start_time = pd.Timestamp('2024-01-08 05:30:00')
        end_time = pd.Timestamp('2024-01-09 23:30:00')
        time_series = pd.date_range(start=start_time, end=end_time, freq='15min')
        num = 73
        target_data1 = target[:num, 0, node_list, 0].numpy()
        zero_data = np.zeros(shape=[23, target_data1.shape[1]])
        target_data2 = target[num:2 * num, 0, node_list, 0].numpy()
        target_data = np.concatenate([target_data1, zero_data, target_data2], axis=0)
        output_data1 = draw_result_Metro(0, num, output, node_list, zero_data)
        output_data2 = draw_result_Metro(1, num, output, node_list, zero_data)
        output_data3 = draw_result_Metro(2, num, output, node_list, zero_data)
        data = np.concatenate([target_data, output_data1, output_data2, output_data3], axis=1)

    else:
        end = 288 * 2
        start_time = pd.Timestamp('2024-01-08 00:00:00')
        end_time = pd.Timestamp('2024-01-09 23:55:00')
        time_series = pd.date_range(start=start_time, end=end_time, freq='5min')
        target_data = target[:end, 0, node_list, 0].numpy()
        output1 = draw_result_Highway(0, end, output, node_list)
        output2 = draw_result_Highway(1, end, output, node_list)
        output3 = draw_result_Highway(2, end, output, node_list)
        data = np.concatenate([target_data, output1, output2, output3], axis=1)

    for i in range(data.shape[1]):
        plt.plot(time_series[2:], data[2:, i], linestyle='-')
    plt.savefig('1.png')
